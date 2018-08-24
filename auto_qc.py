import tkinter as tk
from tkinter import messagebox, filedialog
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
import sys, os, re, subprocess
import traceback
from io import StringIO
from statistics import mean
import xlrd
from simplekml import Kml, Style
import datetime

if getattr(sys, 'frozen', False):
    CWD = sys._MEIPASS
else:
    CWD = os.path.dirname(os.path.realpath(__file__))

QC_PATH = None
TEMPLATE_FILE = os.path.join(CWD, 'template.xlsx')
FWD_TEST_LIST_FILE = None
PROJECT_NAME = ''
#DATE_RE = '_(0[1-9]|1[012])(0[1-9]|[12][0-9]|3[01])[0-9]{2}\.mdb$'
MDB_RE = '\.mdb$'
PYTHON_DIR = ['C:\\Python27', 'C:\\ESRI\\Python2.7', 'C:\\Python2.7', 'C:\\ESRI\\Python27']
FONT_CHECKMARK = Font(name='Wingdings 2', size=11, bold=True)
ARCGIS_PY_SCRIPT = os.path.join(CWD, 'gen_arcgis_file.py')
MAKE_ARCGIS_FILE = False
MDB_FILES = None

GPS_BOUNDS = [{'max_lat':60,'min_lat':49,'min_long':-120,'max_long':-110},\
            {'max_lat':60,'min_lat':49,'min_long':-139.058333,'max_long':-114.058333},\
            {'max_lat':56.9,'min_lat':41.908333,'min_long':-95.153333,'max_long':-74.343333},\
            {'max_lat':70,'min_lat':41.908333,'min_long':-141,'max_long':-52.6166667},\
            {'max_lat':49.388888,'min_lat':24.543333,'min_long':-124.5666667,'max_long':-66.95}]
GPS_CHECK = None

Summary_Stats_List = []
Section_No_Dict = {}
Data_List = []
Station_IDs_Dict = {}
Insufficient_Tests_Dict = {}
Data_Headers = None
Fwd_Test_List = []
Fwd_Test_List_Dict = {}
Stations_Data_List = []
Stations_Data_Headers = None

MAX_AIR_TEMP = None
MIN_AIR_TEMP = None
MAX_SURFACE_TEMP = None
MIN_SURFACE_TEMP = None
DROP_FORCE = [None] * 4

D1_CHECK = None
D1_UNITS = None

tk_root = None
middle_frame = None
tk_chkbox = None
tk_dir_entry = None
tk_dir_entry_str = None
tk_file_entry = None
tk_file_entry_str = None
tk_dir_button = None
tk_file_button = None
tk_run_button = None
tk_max_airtemp_entry = None
tk_min_airtemp_entry = None
tk_max_surtemp_entry = None
tk_min_surtemp_entry = None
tk_name_str = None
tk_name_entry = None
tk_gps_var = None
tk_defl_var = None
tk_defl_units_var = None
tk_force_d1_entry = tk_force_d2_entry = tk_force_d3_entry = tk_force_d4_entry = None
highlight = NamedStyle(name="highlight")
highlight.fill = PatternFill("solid", fgColor="f5b7b1")

class Summary_Stats:
    def __init__(self, mdb_file_name, date, num_drops, num_stations, max_station, min_surface_temp, max_surface_temp, 
        min_air_temp, max_air_temp, station_ids, from_time, to_time, station_check, surface_temp_check, air_temp_check, gps_check):
        self.date = date
        self.max_station = max_station
        self.min_surface_temp = min_surface_temp
        self.max_surface_temp = max_surface_temp
        self.min_air_temp = min_air_temp
        self.max_air_temp = max_air_temp
        self.station_ids = station_ids
        self.mdb_file_name = mdb_file_name
        self.num_drops = num_drops
        self.num_stations = num_stations
        self.sect_no_check = ''
        self.from_time = from_time
        self.to_time = to_time
        self.station_check = station_check
        self.surface_temp_check = surface_temp_check
        self.air_temp_check = air_temp_check
        self.gps_check = gps_check

def is_number(value):
  try:
    float(value)
    return True
  except:
    return False

def left(string, delim):
    pos = string.rfind(delim)
    if pos == -1:
        return string
    return string[:pos]

def is_mdb_file(file):
    m = re.search(MDB_RE, file)
    if m:
        return True
    return False

def find_python27():
    r = 'python.exe'
    for path in PYTHON_DIR:
        if os.path.isdir(path):
            for _path, subFolders, files in os.walk(path):
                for file_name in files:
                    m = re.match(r, file_name)
                    if m:
                        return os.path.join(_path, file_name)

def find_mdb_files(path):
    files_dict = {}
    for _path, subFolders, files in os.walk(path):
        for file_name in files:
            if is_mdb_file(file_name.lower()):
                file_path = os.path.join(_path, file_name)
                files_dict[file_name] = ((file_name, file_path))
    return sorted(files_dict.values())

def add_to_section_dict(mdb_data, mdb_file_name):
    global Section_No_Dict
    section_no_dict = {}
    stationID_col = get_col_no(Data_Headers, 'StationID')
    sect_col = get_col_no(Data_Headers, 'SECT_NO')
    if sect_col != None and stationID_col != None:
        for row in mdb_data:
            sect_no = row[sect_col]
            key = (row[stationID_col], sect_no)
            if key not in section_no_dict:
                if sect_no in Section_No_Dict:
                    count, mdb_file_names_dict = Section_No_Dict[sect_no]
                    count += 1
                    mdb_file_names_dict[mdb_file_name] = None
                    Section_No_Dict[sect_no] = [count, mdb_file_names_dict]
                else:
                    Section_No_Dict[sect_no] = [1, {mdb_file_name: None}]
                section_no_dict[key] = None

def add_sect_no_check(mdb_file_name, sect_no):
    for ss in Summary_Stats_List:
        if ss.mdb_file_name == mdb_file_name:
            if ss.sect_no_check:
                ss.sect_no_check += ', ' + sect_no
            else:
                ss.sect_no_check = 'Sect_No: ' + sect_no
            break

def make_test_list_dict():
    global Fwd_Test_List_Dict
    sect_no_col = get_col_no(Fwd_Test_List[0], ['RT_NO', 'SECT_NO', 'FWD_NO'])
    if sect_no_col != None:
        for row in Fwd_Test_List[1:]:
            sect_no = row[sect_no_col]
            if is_number(sect_no):
                sect_no = str(int(float(sect_no)))
                Fwd_Test_List_Dict[sect_no] = row

def check_test_list_file(file):
    wb = xlrd.open_workbook(file)
    for ws in wb.sheets():
        headers = ws.row(0)
        if headers:
            header_list = []
            for j, col in enumerate(range(ws.ncols)):
                header_list.append(ws.cell_value(0, j))
            sect_no_col = get_col_no(header_list, ['RT_NO', 'SECT_NO', 'FWD_NO'])
            total_tests_col = get_col_no(header_list, ['TOTAL TESTS','TOTAL', 'TOT_TEST'])
            if sect_no_col != None and total_tests_col != None:
                return ws
    return False

def read_fwd_test_list():
    global Fwd_Test_List
    if FWD_TEST_LIST_FILE == None:
        return
    ws = check_test_list_file(FWD_TEST_LIST_FILE)
    if ws != False:
        for i, row in enumerate(range(ws.nrows)):
            r = []
            for j, col in enumerate(range(ws.ncols)):
                r.append(ws.cell_value(i, j))
            Fwd_Test_List.append(r)
        if Fwd_Test_List:
            Fwd_Test_List[0].extend(['Field Tests', 'Compare Test Count', 'Insufficient Field Tests', 'Comments'])
            make_test_list_dict()

def check_section_length(station, section_no):
    if not (Fwd_Test_List_Dict and Fwd_Test_List):
        return
    length_col = get_col_no(Fwd_Test_List[0], 'LENGTH')
    if length_col != None:
        section_no = str(section_no)
        if section_no in Fwd_Test_List_Dict:
            row = Fwd_Test_List_Dict[section_no]
            if length_col < len(row):
                length = row[length_col]
                if is_number(station) and is_number(length):
                    if float(station) > float(length):
                        return str(int(float(length)))

def get_rm_no(section_no):
    if not (Fwd_Test_List_Dict and Fwd_Test_List):
        return
    rm_no_col = get_col_no(Fwd_Test_List[0], 'RM_NO')
    if rm_no_col != None:
        section_no = str(section_no)
        if section_no in Fwd_Test_List_Dict:
            row = Fwd_Test_List_Dict[section_no]
            if rm_no_col < len(row):
                return row[rm_no_col]

def get_col(headers, header_name):
    try:
        return headers.index(header_name)
    except:
        return None

def get_col_no(headers, names):
    if type(names) == str:
        return get_col([str(h).upper() for h in headers], names.upper())
    elif type(names) == list:
        fields = [str(h).upper() for h in headers]
        for name in names:
            index = get_col(fields, name.upper())
            if index != None:
                return index

def style_cell(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = FONT_CHECKMARK
    c.alignment = Alignment(horizontal='center')

def write_test_list_ws(wb):
    global Insufficient_Tests_Dict
    test_list_ws = wb.worksheets[1]
    if 'test list' in wb:
        test_list_ws = wb['test list']

    # write field test numbers, test list sheet
    headers = Fwd_Test_List[0]
    sect_no_col = get_col_no(headers, ['RT_NO', 'SECT_NO', 'FWD_NO'])
    total_tests_col = get_col_no(headers, ['TOTAL TESTS','TOTAL', 'TOT_TEST'])

    if sect_no_col != None and total_tests_col != None:
        for col, header in enumerate(headers, start=1):
            test_list_ws.cell(row=1, column=col).value = header
        for i,row in enumerate(Fwd_Test_List[1:]):
            if len(row) > total_tests_col and sect_no_col < len(row):
                style = None
                sect_no = str(row[sect_no_col])
                if is_number(sect_no):
                    sect_no = str(int(float(sect_no)))
                if sect_no in Section_No_Dict:
                    field_tests, mdb_names_dict = Section_No_Dict[sect_no]
                    reqd_tests = row[total_tests_col]
                    row.append(field_tests)
                    if is_number(reqd_tests):
                        reqd_tests = int(round(float(reqd_tests), 0))
                        check = reqd_tests - field_tests
                        row.append(check)
                        if check > 0:
                            for mdb_file_name in mdb_names_dict.keys():
                                add_sect_no_check(mdb_file_name, sect_no)
                            row.append('O')
                            style = [i + 2, len(row)]
                            Insufficient_Tests_Dict[sect_no] = None
                else:
                    row.append('0')
                    reqd_tests = row[total_tests_col]
                    if is_number(reqd_tests):
                        reqd_tests = int(round(float(reqd_tests), 0))
                        row.append(reqd_tests)
                        if reqd_tests > 0:
                            row.append('O')
                            style = [i + 2, len(row)]
                            Insufficient_Tests_Dict[sect_no] = None
                test_list_ws.append(row)
                if style != None:
                    style_cell(test_list_ws, style[0], style[1])
    else:
        for row in Fwd_Test_List:
            test_list_ws.append(row)

def write_summary_ws(wb):
    summary_ws = wb.worksheets[0]
    if 'summary' in wb:
        summary_ws = wb['summary']

    # write summary stats
    for i,s in enumerate(Summary_Stats_List):
        if not s.sect_no_check and FWD_TEST_LIST_FILE:
            s.sect_no_check = 'P'
        if not s.station_check and FWD_TEST_LIST_FILE:
            s.station_check = 'P'
        row = (s.date, s.from_time, s.to_time, s.mdb_file_name, '', '', '', s.num_stations, s.num_drops, '', s.sect_no_check, s.max_station, s.surface_temp_check,
            s.min_surface_temp, s.max_surface_temp, s.air_temp_check, s.min_air_temp, s.max_air_temp, s.gps_check, s.station_ids or 'P', s.station_check)
        summary_ws.append(row)
        if not s.station_ids:
            style_cell(summary_ws, i + 2, 20)
        if s.sect_no_check == 'P' and FWD_TEST_LIST_FILE:
            style_cell(summary_ws, i + 2, 11)
        if s.station_check == 'P' and FWD_TEST_LIST_FILE:
            style_cell(summary_ws, i + 2, 21)

def in_station_id_dict(row, headers, stn_data_bool=False):
    file_col = get_col_no(headers, 'File')
    stationID_col = get_col_no(headers, 'StationID')
    if stn_data_bool:
        for key in Station_IDs_Dict.keys():
            stn_id, file, drop_id = key
            if stationID_col != None:
                if row[file_col] == file and row[stationID_col] == stn_id:
                    return Station_IDs_Dict[key]
    else:
        drop_col = get_col_no(headers, 'DropID')
        if stationID_col != None and drop_col != None:
            key = (row[stationID_col], row[file_col], row[drop_col])
            if key in Station_IDs_Dict:
                return Station_IDs_Dict[key]
    return False

def highlight_cell(ws, row, col, highlight):
    xls_cell = ws.cell(row=row, column=col)
    xls_cell.style = highlight

def write_data_ws(ws, data_list, headers, stn_data_bool=False):
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    highlight_cell(ws, 1, len(headers) - 1, highlight)
    highlight_cell(ws, 1, len(headers) - 2, highlight)
    if D1_CHECK != None and not stn_data_bool:
        highlight_cell(ws, 1, len(headers) - 3, highlight)
    highlight_cell(ws, 1, len(headers), highlight)
    station_col = get_col_no(headers, 'Station')

    defl_col = get_col_no(headers, 'Decreasing Deflections')
    tests_col = get_col_no(headers, 'Insufficient Field Tests')
    length_col = get_col_no(headers, 'Station < Section Length')
    defl_tol_col = get_col_no(headers, 'Deflection Tolerance')
    drop_force_col = get_col_no(headers, 'Drop Force within Specification')

    for i,row in enumerate(data_list):
        ws.append(row)
        if defl_col != None and len(row[defl_col]) > 0:
            # increasing deflection, highlight this cell
            if not stn_data_bool:
                highlight_cell(ws, i + 2, row[defl_col] + 1, highlight)
            style_cell(ws, i + 2, defl_col + 1)
        if tests_col != None and row[tests_col] == 'O':
            style_cell(ws, i + 2, tests_col + 1)
        if defl_tol_col != None and row[defl_tol_col] == 'O':
            style_cell(ws, i + 2, defl_tol_col + 1)
        if drop_force_col != None and row[drop_force_col] == 'O':
            style_cell(ws, i + 2, drop_force_col + 1)
        if length_col != None and len(row[length_col]) > 0:
            highlight_cell(ws, i + 2, station_col + 1, highlight)
            highlight_cell(ws, i + 2, length_col + 1, highlight)

def write_station_ws(wb):
    # write station & drops data
    if not 'Stations' in wb:
        wb.create_sheet('Stations')
    stations_ws = wb['Stations']
    drops_ws = wb['Stations & Drops']
    global Data_Headers, Data_List, Stations_Data_List, Stations_Data_Headers
    write_data_ws(drops_ws, Data_List, Data_Headers)
    write_data_ws(stations_ws, Stations_Data_List, Stations_Data_Headers, True)

def write_excel_file():
    wb = load_workbook(TEMPLATE_FILE)
    if Fwd_Test_List:
        write_test_list_ws(wb)
    write_summary_ws(wb)
    write_station_ws(wb)
    qc_file = os.path.join(QC_PATH, PROJECT_NAME + '.xlsx')
    wb.save(qc_file)
    return qc_file

def write_kml(file_path, color, row_dict=None):
    sect_col = get_col_no(Stations_Data_Headers, 'SECT_NO')
    slab_col = get_col_no(Stations_Data_Headers, 'SlabID')
    lat_col = get_col_no(Stations_Data_Headers, 'Latitude')
    long_col = get_col_no(Stations_Data_Headers, 'Longitude')

    file_col = get_col_no(Stations_Data_Headers, 'FILE')
    stationID_col = get_col_no(Stations_Data_Headers, 'StationID')
    station_col = get_col_no(Stations_Data_Headers, 'Station')
    lane_col = get_col_no(Stations_Data_Headers, 'Lane')
    if not (lat_col and long_col):
        return

    kml = Kml()
    sharedstyle = Style()
    sharedstyle.iconstyle.icon.href = 'http://www.google.com/intl/en_us/mapfiles/ms/icons/' + color + '-dot.png'
    #'http://maps.google.com/mapfiles/kml/paddle/' + color + '-circle.png'
    for i,row in enumerate(Data_List):
        if not row_dict or  i in row_dict:
            if sect_col != None and slab_col != None and sect_col < len(row) and slab_col < len(row):
                pt = kml.newpoint(name=str(row[slab_col]), coords=[(row[long_col], row[lat_col])])
                if file_col != None and file_col < len(row):
                    pt.extendeddata.newdata(name='File', value=row[file_col])
                pt.extendeddata.newdata(name='Sect_No', value=row[sect_col])
                if stationID_col != None and stationID_col < len(row):
                    pt.extendeddata.newdata(name='StationID', value=row[stationID_col])
                if station_col != None and station_col < len(row):
                    stn = row[station_col]
                    if is_number(stn):
                        stn = round(float(row[station_col]), 1)
                    pt.extendeddata.newdata(name='Station', value=stn)
                if lane_col != None and lane_col < len(row):
                    pt.extendeddata.newdata(name='Lane', value=row[lane_col])
                pt.style = sharedstyle
    kml.save(file_path)

def write_kml_file():
    file_path = os.path.join(QC_PATH, PROJECT_NAME + '.kml')
    write_kml(file_path, 'blue')
    return file_path

def write_bad_drops_kml():
    if not Station_IDs_Dict:
        return
    row_dict = {}
    for i,row in enumerate(Data_List):
        if in_station_id_dict(row, Stations_Data_Headers):
            row_dict[i] = None
    if row_dict:
        write_kml(os.path.join(QC_PATH, PROJECT_NAME + '_stations_with_increasing_deflection.kml'), 'red', row_dict)

def write_bad_sections_kml():
    sect_col = get_col_no(Stations_Data_Headers, 'SECT_NO')
    stationID_col = get_col_no(Stations_Data_Headers, 'StationID')
    if not (Insufficient_Tests_Dict and sect_col and stationID_col):
        return
    row_dict = {}
    section_no_dict = {}
    file_col = get_col_no(Stations_Data_Headers, 'File')
    
    for i,row in enumerate(Data_List):
        key = (row[sect_col], row[stationID_col], row[file_col])
        if key not in section_no_dict and row[sect_col] in Insufficient_Tests_Dict:
            row_dict[i] = None
            section_no_dict[key] = None
    if row_dict:
        write_kml(os.path.join(QC_PATH, PROJECT_NAME + '_sections_with_insufficient_tests.kml'), 'red', row_dict)

def check_drop_force(force, drop_no):
    ret = True
    if drop_no >= 0 and drop_no < len(DROP_FORCE):
        drop_force = DROP_FORCE[drop_no]
        if is_number(force) and is_number(drop_force) and (force > drop_force * 1.1 or force < drop_force * 0.9):
            ret = False
    return ret

def check_coords(lats,longs):
    if not (lats and longs):
        return False
    chk = True
    avg_lat = mean(lats)
    avg_long = mean(longs)
    bounds = None
    tolerance = 4

    if GPS_CHECK != None:
        bounds = GPS_BOUNDS[GPS_CHECK]
    for lat in lats:
        if not lat:
            chk = False
            break
        elif lat > avg_lat + tolerance or lat < avg_lat - tolerance:
            chk = False
            break
        elif bounds != None:
            if lat > bounds['max_lat'] or lat < bounds['min_lat']:
                chk = False
                break
    for lng in longs:
        if not lng:
            chk = False
            break
        elif lng > avg_long + tolerance or lng < avg_long - tolerance:
            chk = False
            break
        elif bounds != None:
            if lng > bounds['max_long'] or lng < bounds['min_long']:
                chk = False
                break
    return chk

def get_sect_info(slabID):
    sect_no = ''
    return_run = ''
    if slabID != None and slabID != '':
        sect_no = left(slabID, '-')
        if len(sect_no) > 1 and sect_no[-1] == '1':
            sect_no = sect_no[:-1] + '0'
            return_run = 'RETURN RUN'
        pos = slabID.find('-')
        if pos != -1:
                slabID = sect_no + slabID[pos:]
        else:
            slabID = sect_no
    return (slabID, sect_no, return_run)

def add_columns(data_rows, mdb_file_name, stn_data_bool=False):
    sect_col = get_col_no(Data_Headers, 'SECT_NO')
    rm_no_col = get_col_no(Data_Headers, 'RM_NO')
    slab_col = get_col_no(Data_Headers, 'SlabID')
    return_run_col = get_col_no(Data_Headers, 'RETURN_RUN')

    mdb_data = []
    for row in data_rows:
        row = [mdb_file_name] + [x for x in row] + ['', '', '', '']
        if not stn_data_bool and D1_CHECK != None:
            row.append('')
        if slab_col != None and len(row) > slab_col and sect_col != None and return_run_col != None:
            slabID, sect_no, return_run = get_sect_info(row[slab_col])
            row[slab_col] = slabID
            row.insert(sect_col, sect_no)
            row.insert(return_run_col, return_run)
            if rm_no_col != None:
                rm_no = get_rm_no(sect_no)
                if rm_no == None:
                    rm_no = ''
                row.insert(rm_no_col, rm_no)
        mdb_data.append(row)
    return mdb_data

def add_headers(headers, stn_data_bool=False):
    headers = ['File'] + [col[0] for col in headers] + ['Decreasing Deflections', 'Insufficient Field Tests', 'Station < Section Length', 'Drop Force within Specification']
    if not stn_data_bool and D1_CHECK != None:
        headers.append('Deflection Tolerance')
    slab_col = get_col_no(headers, 'SlabID')
    if slab_col != None:
        headers.insert(slab_col + 1, 'SECT_NO')
        headers.insert(slab_col + 2, 'RETURN_RUN')
        if FWD_TEST_LIST_FILE and Fwd_Test_List:
            rm_no_col = get_col_no(Fwd_Test_List[0], 'RM_NO')
            if rm_no_col != None:
                headers.insert(slab_col + 3, 'RM_NO')
    return headers

def add_checks(data_list, headers, stn_data_bool=False):
    sect_col = get_col_no(headers, 'SECT_NO')
    station_col = get_col_no(headers, 'Station')
    d1_col = get_col_no(headers, 'D1')
    defl_col = get_col_no(headers, 'Decreasing Deflections')
    tests_col = get_col_no(headers, 'Insufficient Field Tests')
    length_col = get_col_no(headers, 'Station < Section Length')
    defl_tol_col = get_col_no(headers, 'Deflection Tolerance')
    drop_force_col = get_col_no(headers, 'Drop Force within Specification')
    
    for row in data_list:
        if station_col != None and sect_col != None and length_col != None:
            stn_chk = check_section_length(row[station_col], row[sect_col])
            if stn_chk != None and type(stn_chk) == str:
                row[length_col] = 'Section Length: ' + stn_chk
        incr_col = in_station_id_dict(row, headers, stn_data_bool)
        if col != False and defl_col != None:
            # increasing deflection
            row[defl_col] = incr_col
        if sect_col != None and tests_col != None and row[sect_col] in Insufficient_Tests_Dict:
            row[tests_col] = 'O'
        if D1_CHECK != None and defl_tol_col != None and not stn_data_bool and d1_col != None and d1_col < len(row):
            d1 = row[d1_col]
            max_d1 = D1_CHECK
            if D1_UNITS == 1:
                max_d1 *= 25.4
            if d1 > max_d1:
                row[defl_tol_col] = 'O'

    if D1_CHECK != None:
        stationid_col = get_col_no(headers, 'StationID')
        force_col = get_col_no(headers, 'Force')
        if stationid_col != None and force_col != None and stationid_col < len(row) and force_col < len(row):
            i = 0
            stn_id = None
            for row in data_list:
                new_stn_id = row[stationid_col]
                if stn_id != new_stn_id:
                    i=0
                else:
                    i += 1
                stn_id = new_stn_id
                d1_check = check_drop_force(row[force_col], i)
                if d1_check == False:
                    row[drop_force_col] = 'O'


def process_mdb_data(mdb_file_name, data_rows, data_headers, stn_rows, stn_headers):
    global Data_Headers, Data_List, Stations_Data_List, Stations_Data_Headers
    if not Data_Headers:
        Data_Headers = add_headers(data_headers)
        Stations_Data_Headers = add_headers(stn_headers, stn_data_bool=True)
    mdb_data = add_columns(data_rows, mdb_file_name)
    Data_List.extend(mdb_data)
    stn_data = add_columns(stn_rows, mdb_file_name, stn_data_bool=True)
    Stations_Data_List.extend(stn_data)

    data_transposed = [*zip(*mdb_data)]
    #summary stats
    num_drops = len(mdb_data)
    num_stations = len(stn_data)

    station_col = get_col_no(Data_Headers, 'Station')
    if station_col != None:
        max_station = round(max(data_transposed[station_col]), 1)
    else:
        max_station = ''
    temp_col = get_col_no(Data_Headers, ['Surface', 'SurfaceTemperature'])
    surface_temp_check = ''
    air_temp_check = ''
    if temp_col != None:
        min_surface_temp = round(min(data_transposed[temp_col]), 1)
        max_surface_temp = round(max(data_transposed[temp_col]), 1)
        if MAX_SURFACE_TEMP and MIN_SURFACE_TEMP:
            if max_surface_temp <= MAX_SURFACE_TEMP and min_surface_temp >= MIN_SURFACE_TEMP:
                surface_temp_check = 'good'
            else:
                surface_temp_check = 'bad'
    else:
        min_surface_temp = ''
        max_surface_temp = ''
    temp_col = get_col_no(Data_Headers, ['Air', 'AirTemperature'])
    if temp_col != None:
        min_air_temp = round(min(data_transposed[temp_col]), 1)
        max_air_temp = round(max(data_transposed[temp_col]), 1)
        if MAX_AIR_TEMP and MIN_AIR_TEMP:
            if max_air_temp <= MAX_AIR_TEMP and min_air_temp >= MIN_AIR_TEMP:
                air_temp_check = 'good'
            else:
                air_temp_check = 'bad'
    else:
        min_air_temp = ''
        max_air_temp = ''
    time_col = get_col_no(Data_Headers, 'Time')
    if time_col != None:
        times = data_transposed[time_col]
        date = times[0].strftime('%m/%d/%Y')
        from_time = min(times).strftime('%H:%M')
        to_time = max(times).strftime('%H:%M')
    else:
        date = ''
        from_time = ''
        to_time = ''

    station_ids = {}
    d1_col = get_col_no(Data_Headers, 'D1')
    drop_col = get_col_no(Data_Headers, 'DropID')
    stationID_col = get_col_no(Data_Headers, 'StationID')
    sect_col = get_col_no(Data_Headers, 'SECT_NO')
    #stn_len_chk_col = get_col_no(Data_Headers, 'Station < Section Length')
    #deflections_chk_col = get_col_no(Data_Headers, 'Decreasing Deflections')
    stn_chk_ids = {}
    for row in mdb_data:
        if station_col != None and sect_col != None and stationID_col != None:
            stn_chk = check_section_length(row[station_col], row[sect_col])
            if stn_chk:# and stn_len_chk_col and stn_len_chk_col < len(row):
                # highlight bad
                stn_chk_ids[str(row[stationID_col])] = None
                #row[stn_len_chk_col] = 'O'
        if d1_col != None and stationID_col != None and drop_col != None:
            for i in range(d1_col, d1_col + 7):
                if row[i] < row[i + 1]:
                    station_ids[str(row[stationID_col])] = None
                    key = (row[stationID_col], mdb_file_name, row[drop_col])
                    Station_IDs_Dict[key] = i + 1
                    #if deflections_chk_col and deflections_chk_col < len(row):
                    #    row[deflections_chk_col] = 'O'
                    break
    if station_ids:
        station_ids = 'StationID ' + ', '.join(station_ids.keys())
    else:
        station_ids = ''
    if stn_chk_ids:
        station_check = 'StationID ' + ', '.join(stn_chk_ids.keys())
    else:
        station_check = ''
    gps_check = ''
    if GPS_CHECK != None:
        lat_col = get_col_no(Data_Headers, 'Latitude')
        long_col = get_col_no(Data_Headers, 'Longitude')
        if lat_col != None and long_col != None and lat_col < len(data_transposed) and long_col < len(data_transposed):
            gps_check = check_coords(data_transposed[lat_col], data_transposed[long_col])
            if gps_check:
                gps_check = 'good'
            else:
                gps_check = 'bad'

    summary_stats = Summary_Stats(mdb_file_name, date, num_drops, num_stations, max_station, min_surface_temp, max_surface_temp, 
        min_air_temp, max_air_temp, station_ids, from_time, to_time, station_check, surface_temp_check, air_temp_check, gps_check)

    global Summary_Stats_List
    Summary_Stats_List.append(summary_stats)
    add_to_section_dict(mdb_data, mdb_file_name)
    add_checks(mdb_data, Data_Headers, stn_data_bool=False)
    add_checks(stn_data, Stations_Data_Headers, stn_data_bool=False)

def query_mdb_data():
    DRV = '{Microsoft Access Driver (*.mdb)}'
    PWD = 'pw'
    for mdb in MDB_FILES:
        mdb_file_name = mdb[0]
        mdb_file_path = mdb[1]
        
        # connect to db
        connection = pyodbc.connect('DRIVER={};DBQ={};PWD={}'.format(DRV, mdb_file_path, PWD))
        cursor = connection.cursor()

        data_query = 'SELECT * FROM Stations Inner Join Drops On Stations.StationID = Drops.StationID'
        data_rows = cursor.execute(data_query).fetchall()
        if len(data_rows) > 0:
            data_headers = cursor.description

            data_query = 'SELECT * FROM Stations'
            stn_rows = cursor.execute(data_query).fetchall()
            stn_headers = cursor.description
            
            cursor.close()
            connection.close()
            print(mdb_file_name)
            process_mdb_data(mdb_file_name, data_rows, data_headers, stn_rows, stn_headers)

def check_selected_dir(wd):
    if wd:
        wd = wd.replace('/', '\\')
        if os.path.isdir(wd):
            mdb_files = find_mdb_files(wd)
            if mdb_files:
                return wd
    return False

def set_file_name_input(proj_name):
    global tk_file_entry_str
    name = left(proj_name, '.') + '_raw_qc'
    tk_dir_entry_str.set(name)

def set_selected_dir():
    global tk_dir_entry_str
    options = {'title':'Select the QC Project Folder'}
    wd = filedialog.askdirectory(**options)
    if wd:
        wd = check_selected_dir(wd)
        if wd:
            tk_dir_entry_str.set(wd)
            #set_file_name_input(proj_name)
        else:
            messagebox.showinfo('', 'Please select the project folder containing ALL of the .mdb files.')

def check_selected_file(wf):
    if wf:
        wf = wf.replace('/', '\\')
        if os.path.isfile(wf) and \
        (wf.lower().endswith('.xls') or wf.lower().endswith('.xlsx') or wf.lower().endswith('.xlsm')) and \
        check_test_list_file(wf):
            return wf
    return False

def set_selected_file():
    global tk_file_entry_str
    ftypes = [('Excel files', '*.xls;*.xlsx;*.xlsm;')]
    options = {'title':'Select the QC Specification Excel File', 'filetypes':ftypes}

    wd = tk_dir_entry_str.get()
    field_data = check_selected_dir(wd)
    if field_data != False:
        phase = left(field_data,'field_data')
        options['initialdir'] = phase
        if len(field_data) > len(phase):
            specification = os.path.join(phase,'specification')
            if os.path.isdir(specification):
                options['initialdir'] = specification
                fwd = os.path.join(specification, 'fwd')
                if os.path.isdir(fwd):
                    options['initialdir'] = fwd
                elif os.path.isdir(os.path.join(specification, 'fwd_setup')):
                    options['initialdir'] = os.path.join(specification, 'fwd_setup')

    wf = filedialog.askopenfilename(**options)
    if wf:
        wf = check_selected_file(wf)
        if wf:
            tk_file_entry_str.set(wf)
        else:
            messagebox.showinfo('', 'This program can only read Excel Test Specification files.')

def set_global_vars():
    global tk_root, QC_PATH, FWD_TEST_LIST_FILE, PROJECT_NAME, MAKE_ARCGIS_FILE, MDB_FILES, \
    MAX_AIR_TEMP, MIN_AIR_TEMP, MAX_SURFACE_TEMP, MIN_SURFACE_TEMP, GPS_CHECK, D1_CHECK, D1_UNITS, \
    DROP_FORCE
    error = False

    path = check_selected_dir(tk_dir_entry_str.get())
    test_list_file = tk_file_entry_str.get()
    if test_list_file:
        test_list_file = check_selected_file(test_list_file)
        if test_list_file == False:
            messagebox.showinfo('', 'This program can only read Excel Test Specification files.')
            error = True

    if path == False:
        messagebox.showinfo('', 'Please select the project folder containing ALL of the .mdb files.')
        error = True

    max_airtemp = tk_max_airtemp_entry.get()
    if max_airtemp != '' and not is_number(max_airtemp):
        tk_max_airtemp_entry.config(bg='LightPink1')
        error = True
    min_airtemp = tk_min_airtemp_entry.get()
    if min_airtemp != '' and not is_number(min_airtemp):
        tk_min_airtemp_entry.config(bg='LightPink1')
        error = True
    max_surtemp = tk_max_surtemp_entry.get()
    if max_surtemp != '' and not is_number(max_surtemp):
        tk_max_surtemp_entry.config(bg='LightPink1')
        error = True
    min_surtemp = tk_min_surtemp_entry.get()
    if min_surtemp != '' and not is_number(min_surtemp):
        tk_min_surtemp_entry.config(bg='LightPink1')
        error = True
    if max_airtemp != '' or min_airtemp != '':
        if max_airtemp == '':
            error = True
            tk_max_airtemp_entry.config(bg='LightPink1')
        elif min_airtemp == '':
            error = True
            tk_min_airtemp_entry.config(bg='LightPink1')
        elif is_number(max_airtemp) and is_number(min_airtemp) and float(max_airtemp) <= float(min_airtemp):
            error = True
            tk_max_airtemp_entry.config(bg='LightPink1')
            tk_min_airtemp_entry.config(bg='LightPink1')
    if max_surtemp != '' or min_surtemp != '':
        if max_surtemp == '':
            error = True
            tk_max_surtemp_entry.config(bg='LightPink1')
        elif min_surtemp == '':
            error = True
            tk_min_surtemp_entry.config(bg='LightPink1')
        elif is_number(max_surtemp) and is_number(min_surtemp) and float(max_surtemp) <= float(min_surtemp):
            error = True
            tk_max_surtemp_entry.config(bg='LightPink1')
            tk_min_surtemp_entry.config(bg='LightPink1')

    force_d1 = tk_force_d1_entry.get()
    force_d2 = tk_force_d2_entry.get()
    force_d3 = tk_force_d3_entry.get()
    force_d4 = tk_force_d4_entry.get()
    if force_d1 != '' and not is_number(force_d1):
        tk_force_d1_entry.config(bg='LightPink1')
        error = True
    if force_d2 != '' and not is_number(force_d2):
        tk_force_d2_entry.config(bg='LightPink1')
        error = True
    if force_d3 != '' and not is_number(force_d3):
        tk_force_d3_entry.config(bg='LightPink1')
        error = True
    if force_d4 != '' and not is_number(force_d4):
        tk_force_d4_entry.config(bg='LightPink1')
        error = True

    if error == False:
        tk_chkbox.config(state='disable')
        tk_dir_entry.config(state='readonly')
        tk_file_entry.config(state='readonly')
        tk_dir_button.config(state='disable')
        tk_file_button.config(state='disable')
        tk_run_button.config(state='disable')
        tk_name_entry.config(state='readonly')
        tk_max_airtemp_entry.config(state='readonly')
        tk_min_airtemp_entry.config(state='readonly')
        tk_max_surtemp_entry.config(state='readonly')
        tk_min_surtemp_entry.config(state='readonly')
        for child in middle_frame.winfo_children():
            if child.winfo_class() == 'Radiobutton':
                child['state'] = 'disabled'

        if max_airtemp != '':
            MAX_AIR_TEMP = float(max_airtemp)
        if min_airtemp != '':
            MIN_AIR_TEMP = float(min_airtemp)
        if max_surtemp != '':
            MAX_SURFACE_TEMP = float(max_surtemp)
        if min_surtemp != '':
            MIN_SURFACE_TEMP = float(min_surtemp)

        if force_d1 != '':
            DROP_FORCE[0] = float(force_d1)
        if force_d2 != '':
            DROP_FORCE[1] = float(force_d2)
        if force_d3 != '':
            DROP_FORCE[2] = float(force_d3)
        if force_d4 != '':
            DROP_FORCE[3] = float(force_d4)

        QC_PATH = path
        if test_list_file:
            FWD_TEST_LIST_FILE = test_list_file
        MAKE_ARCGIS_FILE = bool(tk_chkbox.var.get())
        mdb_files = find_mdb_files(path)
        proj = mdb_files[0][0]

        gps_check = tk_gps_var.get()
        if gps_check != -1:
            GPS_CHECK = gps_check

        defl_check = tk_defl_var.get()
        if defl_check != -1:
            D1_CHECK = defl_check
            D1_UNITS = tk_defl_units_var.get()

        proj_name = tk_name_entry.get()
        if proj_name != '':
            PROJECT_NAME = proj_name.lower() + '_' + datetime.datetime.now().strftime('%Y%m%d')
        else:
            PROJECT_NAME = left(proj, '.').lower() + '_raw_qc_' + datetime.datetime.now().strftime('%Y%m%d')
        MDB_FILES = mdb_files

        main()
        #tk_root.destroy()
        #tk_root.withdraw()

def open_qc_file(qc_file):
    excel_p = subprocess.Popen('"' + qc_file + '"', shell=True)

def to_cmd_str(args):
    return ('"{}" '*len(args)).format(*args).rstrip()

def make_arcgis_shape_file(kml_file):
    global MAKE_ARCGIS_FILE
    if MAKE_ARCGIS_FILE:
        print('\nMaking ArcGIS layer with tests\n')
        py_exe = find_python27()
        if py_exe:
            cmd = to_cmd_str([py_exe, ARCGIS_PY_SCRIPT, kml_file, QC_PATH])
            try:
                ret = subprocess.check_call(cmd, shell=True)
            except Exception as e:
                handle_exception(e)

def open_files(qc_file):
    open_qc_file(qc_file)
    subprocess.Popen('explorer ' + QC_PATH)

def main():
    if not (QC_PATH and MDB_FILES):
        exit()
    read_fwd_test_list()
    query_mdb_data()
    qc_file = write_excel_file()
    kml_file = write_kml_file()
    write_bad_sections_kml()
    write_bad_drops_kml()
    make_arcgis_shape_file(kml_file)
    open_files(qc_file)
    print('Success!')

def set_up_gui():
    global tk_root, tk_chkbox, tk_dir_entry, tk_file_entry, tk_file_entry_str, tk_dir_entry_str, tk_dir_button, tk_file_button, tk_run_button, \
    tk_max_airtemp_entry, tk_min_airtemp_entry, tk_max_surtemp_entry, tk_min_surtemp_entry, tk_name_str, tk_name_entry, tk_gps_var, tk_defl_var, \
    tk_force_d1_entry, tk_force_d2_entry, tk_force_d3_entry, tk_force_d4_entry, tk_defl_units_var, middle_frame
    tk_root = tk.Tk()
    frame = tk.Frame(tk_root, borderwidth=1)
    frame.pack(side=tk.TOP)
    middle_frame = tk.Frame(tk_root, borderwidth=10)
    middle_frame.pack()
    mid_frame = tk.Frame(tk_root, borderwidth=2)
    mid_frame.pack()
    bottom_frame = tk.Frame(tk_root, borderwidth=2)
    bottom_frame.pack(side=tk.BOTTOM)

    tk_root.wm_title('Auto QC')
    tk.Label(frame, text='Select the Project Folder Containing .mdb files', 
        font = "Helvetica 18").grid(row=0, column=1, columnspan=3, padx=15,pady=15)
    tk.Label(frame, text='Select the FWD Test List File', 
        font = "Helvetica 18").grid(row=2, column=1, columnspan=3, padx=15,pady=15)
    #tk_chkbox.select()
    tk.Label(bottom_frame, text='  \u2794 ', font = "Helvetica 14").grid(row=10, column=6, columnspan=1)
    tk_dir_button = tk.Button(frame, text='Browse', font='Helvetica 12', command=set_selected_dir)
    tk_dir_button.grid(row=1, column=6, sticky=tk.W, padx=15,pady=5)
    tk_file_button = tk.Button(frame, text='Browse', font='Helvetica 12', command=set_selected_file)
    tk_file_button.grid(row=3, column=6, sticky=tk.W, padx=15,pady=5)
        
    tk.Label(frame, text='Project Folder:', 
        font = "Helvetica 12").grid(row=1, column=0, columnspan=1, padx=5,pady=5)
    tk.Label(frame, text='FWD Test List File:\n(optional)', 
        font = "Helvetica 12").grid(row=3, column=0, columnspan=1, padx=5,pady=5)
    tk_dir_entry_str = tk.StringVar()
    tk_dir_entry = tk.Entry(frame, textvariable=tk_dir_entry_str, width=100, readonlybackground='grey85')
    tk_dir_entry.grid(row=1, column=1, columnspan=5)
    tk_file_entry_str = tk.StringVar()
    tk_file_entry = tk.Entry(frame, textvariable=tk_file_entry_str, width=100, readonlybackground='grey85')
    tk_file_entry.grid(row=3, column=1, columnspan=5)

    ###  temperature inputs: (need a separate frame)
    tk.Label(bottom_frame, text='Issues, questions: Eric.Rothfels@stantec.com', 
        font = "Helvetica 11").grid(row=10, column=0, columnspan=4, padx=5,pady=5)
    tk.Label(bottom_frame, text='Air Temperature:\n(optional)', 
        font = "Helvetica 11").grid(row=0, column=1, columnspan=1, padx=5,pady=5)
    tk.Label(bottom_frame, text='Surface Temperature:\n(optional)', 
        font = "Helvetica 11").grid(row=1, column=1, columnspan=1, padx=5,pady=5)
    tk.Label(bottom_frame, text='Min:', 
        font = "Helvetica 11").grid(row=0, column=2, columnspan=1, padx=5,pady=5)
    tk.Label(bottom_frame, text='Max:', 
        font = "Helvetica 11").grid(row=0, column=4, columnspan=1, padx=5,pady=5)
    tk.Label(bottom_frame, text='Min:', 
        font = "Helvetica 11").grid(row=1, column=2, columnspan=1, padx=5,pady=5)
    tk.Label(bottom_frame, text='Max:', 
        font = "Helvetica 11").grid(row=1, column=4, columnspan=1, padx=5,pady=5)
    tk_min_airtemp_str = tk.StringVar()
    tk_min_airtemp_entry = tk.Entry(bottom_frame, textvariable=tk_min_airtemp_str, width=10, readonlybackground='grey85')
    tk_min_airtemp_entry.grid(row=0, column=3, columnspan=1)
    tk_max_airtemp_str = tk.StringVar()
    tk_max_airtemp_entry = tk.Entry(bottom_frame, textvariable=tk_max_airtemp_str, width=10, readonlybackground='grey85')
    tk_max_airtemp_entry.grid(row=0, column=5, columnspan=1)
    tk_min_surtemp_str = tk.StringVar()
    tk_min_surtemp_entry = tk.Entry(bottom_frame, textvariable=tk_min_surtemp_str, width=10, readonlybackground='grey85')
    tk_min_surtemp_entry.grid(row=1, column=3, columnspan=1)
    tk_max_surtemp_str = tk.StringVar()
    tk_max_surtemp_entry = tk.Entry(bottom_frame, textvariable=tk_max_surtemp_str, width=10, readonlybackground='grey85')
    tk_max_surtemp_entry.grid(row=1, column=5, columnspan=1)
    tk.Label(bottom_frame, text='Output File Name:\n(optional)', 
        font = "Helvetica 11").grid(row=2, column=1, columnspan=1, padx=5,pady=5)
    tk_name_str = tk.StringVar()
    tk_name_entry = tk.Entry(bottom_frame, textvariable=tk_name_str, width=30, readonlybackground='grey85')
    tk_name_entry.grid(row=2, column=2, columnspan=3)
    v = tk.IntVar()
    tk_chkbox = tk.Checkbutton(bottom_frame, text="Make ArcGIS Shape File from FWD Tests", variable=v)
    tk_chkbox.var = v
    tk_chkbox.grid(row=3, column=2, columnspan=3, padx=15,pady=20)
    tk_run_button = tk.Button(bottom_frame, text='Run', font='Helvetica 14', command=set_global_vars)
    tk_run_button.grid(row=10, column=7, sticky=tk.W, padx=15,pady=15)

    tk.Label(middle_frame, text='Location of FWD Tests:\nfor GPS Check\n(optional)', 
        font = "Helvetica 11").grid(row=0, column=0, rowspan=3,columnspan=1, padx=5,pady=5)
    tk_gps_var = tk.IntVar()
    R1 = tk.Radiobutton(middle_frame, text="No GPS Check", variable=tk_gps_var, value=-1)
    R1.grid(row=0, column=1)
    R1.select()
    R1 = tk.Radiobutton(middle_frame, text="Alberta", variable=tk_gps_var, value=0)
    R1.grid(row=1, column=1)
    R2 = tk.Radiobutton(middle_frame, text="British Columbia", variable=tk_gps_var, value=1)
    R2.grid(row=2, column=1)
    R3 = tk.Radiobutton(middle_frame, text="Ontario", variable=tk_gps_var, value=2)
    R3.grid(row=0, column=2)
    R1 = tk.Radiobutton(middle_frame, text="Canada", variable=tk_gps_var, value=3)
    R1.grid(row=1, column=2)
    R2 = tk.Radiobutton(middle_frame, text="USA (Contiguous)", variable=tk_gps_var, value=4)
    R2.grid(row=2, column=2)
    tk.Label(middle_frame, text=' ').grid(row=3, column=0, rowspan=3,columnspan=1, padx=5,pady=2)
    
    tk_defl_var = tk.IntVar()
    tk.Label(middle_frame, text='Deflection Tolerance Check:\n(optional)', 
        font = "Helvetica 11").grid(row=7, column=0, rowspan=3,columnspan=1, padx=5,pady=5)
    R1 = tk.Radiobutton(middle_frame, text="No Deflection Tolerance Check", variable=tk_defl_var, value=-1)
    R1.grid(row=7, column=1)
    R1.select()
    R1 = tk.Radiobutton(middle_frame, text="FWD", variable=tk_defl_var, value=80)
    R1.grid(row=8, column=1)
    R2 = tk.Radiobutton(middle_frame, text="HWD", variable=tk_defl_var, value=100)
    R2.grid(row=9, column=1)
    tk_defl_units_var = tk.IntVar()
    R2 = tk.Radiobutton(middle_frame, text="Microns (micrometers)", variable=tk_defl_units_var, value=1)
    R2.grid(row=8, column=2)
    R2.select()
    R1 = tk.Radiobutton(middle_frame, text="Mils (milli inches)", variable=tk_defl_units_var, value=0)
    R1.grid(row=9, column=2)

    tk.Label(mid_frame, text=' ').grid(row=9, column=0, rowspan=3,columnspan=1, padx=5,pady=2)
    tk.Label(mid_frame, text='Specified Drop Force:\n(optional)', 
        font = "Helvetica 11").grid(row=10, column=1, columnspan=1, padx=5,pady=5)
    tk.Label(mid_frame, text='Drop 1:', 
        font = "Helvetica 11").grid(row=10, column=2, columnspan=1, padx=5,pady=5)
    tk.Label(mid_frame, text='Drop 2:', 
        font = "Helvetica 11").grid(row=10, column=4, columnspan=1, padx=5,pady=5)
    tk.Label(mid_frame, text='Drop 3:', 
        font = "Helvetica 11").grid(row=10, column=6, columnspan=1, padx=5,pady=5)
    tk.Label(mid_frame, text='Drop 4:', 
        font = "Helvetica 11").grid(row=10, column=8, columnspan=1, padx=5,pady=5)
    tk_force_d1_str = tk.StringVar()
    tk_force_d1_entry = tk.Entry(mid_frame, textvariable=tk_force_d1_str, width=6, readonlybackground='grey85')
    tk_force_d1_entry.grid(row=10, column=3, columnspan=1)
    tk_force_d2_str = tk.StringVar()
    tk_force_d2_entry = tk.Entry(mid_frame, textvariable=tk_force_d2_str, width=6, readonlybackground='grey85')
    tk_force_d2_entry.grid(row=10, column=5, columnspan=1)
    tk_force_d3_str = tk.StringVar()
    tk_force_d3_entry = tk.Entry(mid_frame, textvariable=tk_force_d3_str, width=6, readonlybackground='grey85')
    tk_force_d3_entry.grid(row=10, column=7, columnspan=1)
    tk_force_d4_str = tk.StringVar()
    tk_force_d4_entry = tk.Entry(mid_frame, textvariable=tk_force_d4_str, width=6, readonlybackground='grey85')
    tk_force_d4_entry.grid(row=10, column=9, columnspan=1)
    tk.Label(mid_frame, text=' ').grid(row=11, column=0, rowspan=3,columnspan=1, padx=5,pady=2)
    
    
    #centre the window
    tk_root.eval('tk::PlaceWindow %s center' % tk_root.winfo_pathname(tk_root.winfo_id()))
    tk_root.mainloop()

def get_exception_stack():
    old_stderr = sys.stderr
    sys.stderr = mystderr = StringIO()
    traceback.print_exc()
    sys.stderr = old_stderr
    return mystderr.getvalue()

def handle_exception(e):
    error = get_exception_stack()
    if DEV_MODE:
        print(error, file=sys.stderr)
    else:
        file = None
        #send_email(error, file)
    error += '\n\nPress OK to Continue, or Cancel to Quit the Program'
    if not messagebox.askokcancel('Error', error, icon=messagebox.ERROR):
        os._exit(0)

if __name__=="__main__":
    print("""\nWelcome to Auto QC \n\nClose this window at any time to Quit the Program.\n
    \n""")
    set_up_gui()
